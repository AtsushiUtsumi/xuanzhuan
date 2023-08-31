from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

# 新しいワークブックを作成
workbook = Workbook()
sheet = workbook.active

# セルの幅を設定
column_widths = [15, 30, 30, 15]
for col, width in enumerate(column_widths, start=1):
    sheet.column_dimensions[chr(64 + col)].width = width

# 方眼紙スタイルの設定
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

# セルにデータとスタイルを設定
header_data = ['項目', '説明', '備考', '優先度']
data = [
    ['機能A', 'これは機能Aの説明です', '', '高'],
    ['機能B', 'これは機能Bの説明です', '特記事項', '中'],
    # 他のデータ行を追加
]

# ヘッダ行を作成
sheet.append(header_data)

# データ行を作成
for row_data in data:
    sheet.append(row_data)

# セルにスタイルを適用
for row in sheet.iter_rows(min_row=1, max_row=1):
    for cell in row:
        cell.border = thin_border
        cell.alignment = alignment
        cell.font = cell.font.copy(bold=True)

for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
    for cell in row:
        cell.border = thin_border
        cell.alignment = alignment
    sheet.row_dimensions[row_index].height = 20  # 行の高さを設定

# ワークブックを保存
workbook.save('仕様書.xlsx')
