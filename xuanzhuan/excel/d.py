from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

# 新しいワークブックを作成
workbook = Workbook()
sheet = workbook.active

for i in range(1,11):
    for j in range(1,11):
        sheet.cell(row=i, column=j,value=f'{i}_{j}')
    sheet.merge_cells(start_row= i, end_row=i, start_column=1, end_column=2)
    sheet.merge_cells(start_row= i, end_row=i, start_column=3, end_column=5)
    sheet.merge_cells(start_row= i, end_row=i, start_column=6, end_column=10)

# ワークブックを保存
workbook.save('セルにインデックスでアクセス.xlsx')
