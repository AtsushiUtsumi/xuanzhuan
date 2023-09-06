from openpyxl import Workbook
from openpyxl.styles import Alignment

import openpyxl
from openpyxl.styles.borders import Border, Side


def create_docs(docs_name: str):
    """
    仕様書のエクセルファイルを出力します

    Args:
        docs_name (string): 出力するファイルの名前(拡張子無し)
    """

    # 新しいワークブックを作成
    workbook = Workbook()
    sheet = workbook.active

    # セルに改行を含むテキストを設定
    text_with_newline = "行1\n行2\n行3"
    sheet["A1"] = text_with_newline

    # セル内のテキストを中央揃えで改行
    alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    sheet["A1"].alignment = alignment

    # ワークブックを保存
    workbook.save(f"{docs_name}.xlsx")
    return


class Docs:
    def __init__(self, docs_name: str) -> None:
        self.cursor = 0  # 現在処理中の行番号
        self.docs_name = docs_name
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        # 列幅はここで設定
        for i in range(1, 50):
            self.sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3
        # self.sheet.column_dimensions['G'].width = 70
        # self.sheet.column_dimensions['K'].width = 50
        pass

    def add_line(self, csv: str):
        self.cursor += 1
        columns = csv.split(",")
        n = max([1 + i.count("\n") for i in columns])  # 最大行数計算
        tmp = [0]
        for c in columns:
            tmp.append(max([len(s) for s in c.split("\n")]))
        max_str_length = max(tmp)
        print(str(n) + "行が最大行数です")
        print(str(max_str_length) + "文字が最大文字数です")
        self.sheet.row_dimensions[self.cursor].height = 14 * n  # 行の高さを設定
        for j in range(len(columns)):
            self.sheet.cell(row=self.cursor, column=j + 1, value=f"{columns[j]}")
            cell = self.sheet.cell(row=self.cursor, column=j + 1, value=f"{columns[j]}")
        alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        for row in self.sheet.iter_rows(min_row=1, max_row=self.cursor):
            for cell in row:
                cell.alignment = alignment
        # エクセル方眼紙だとセル結合が必要だからめんどい
        # 番号, テスト区分, 前提, 操作, 期待される結果, 結果,名前,日付
        self.sheet.merge_cells(
            start_row=self.cursor, end_row=self.cursor, start_column=1, end_column=11
        )
        self.sheet.merge_cells(
            start_row=self.cursor, end_row=self.cursor, start_column=12, end_column=22
        )
        self.sheet.merge_cells(
            start_row=self.cursor, end_row=self.cursor, start_column=23, end_column=33
        )
        return

    def save(self):
        side = Side(style='thin', color='000000')
        #side = Side(style='medium', color='000000')
        #side = Side(style='thick', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        for row in self.sheet:
            for cell in row:
                self.sheet[cell.coordinate].border = border
        self.workbook.save(f"{self.docs_name}.xlsx")
        return

    def add_merged_line(self, csv: str):# 「"A:F:hoge, G:I:fuga, J:K:hage"」ときたらA～Fカラムに「hoge」,G～Iカラムに「fuga」,J～Kカラムに「hage」としたい
        # TODO: 帰ったらこれ実装する
        return
